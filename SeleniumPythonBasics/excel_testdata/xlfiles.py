import openpyxl

path_of_xl = r"C:\Users\user\Desktop\sample.xlsx"
# workbook--> sheet--> row&col--> cells

# read
wb = openpyxl.load_workbook(path_of_xl)
#print(wb.sheetnames)# ['Sheet1', 'Sheet2', 'Sheet3']
#print(wb.active)
sheet = wb["Sheet1"]
print("No of rows", sheet.max_row)
print("No of col", sheet.max_column)
row=sheet.max_row
col=sheet.max_column

# python range will start from 0, n-1
for r in range(1, row+1):
    for c in range(1,col+1):
        print(sheet.cell(r,c).value, end= "\t")
    print("\n")

