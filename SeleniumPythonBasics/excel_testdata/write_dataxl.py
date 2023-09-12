import openpyxl

path_of_xl = r"C:\Users\user\Desktop\sample.xlsx"
# workbook--> sheet--> row&col--> cells
wb = openpyxl.load_workbook(path_of_xl)
sheet=wb["Sheet3"]

sheet.cell(1,1).value="student1"
sheet.cell(1,2).value="1234"

sheet.cell(2,1).value="student2"
sheet.cell(2,2).value="1234556"

for r in range(1,sheet.max_row+1):
    sheet.cell(r, 3).value = "NA"

wb.save(path_of_xl)