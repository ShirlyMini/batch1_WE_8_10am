import time
import openpyxl
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.select import Select
path_of_xl = r"C:\Users\user\Desktop\simple_interest.xlsx"
wb=openpyxl.load_workbook(path_of_xl)
sheet=wb["Sheet1"]
print("No of rows", sheet.max_row)
print("No of col", sheet.max_column)


driver = webdriver.Chrome()
driver.get("https://cleartax.in/s/simple-compound-interest-calculator")
driver.maximize_window()

for r in range(2, sheet.max_row+1):

    principal=sheet.cell(r,1).value
    rate=sheet.cell(r,2).value
    period_op=sheet.cell(r,3).value
    period_unit=sheet.cell(r,4).value
    int_type=sheet.cell(r,5).value
    total_value = sheet.cell(r,6).value

    drpdown_elem=driver.find_element(By.ID,"a")
    type_drpdown=Select(drpdown_elem)
    type_drpdown.select_by_visible_text(int_type)

    driver.find_element(By.ID,"c").clear()
    driver.find_element(By.ID,"c").send_keys(principal)
    driver.find_element(By.ID,"d").clear()
    driver.find_element(By.ID,"d").send_keys(rate)

    period_dropdown=driver.find_element(By.ID,"f")
    period_unit_drpdown=Select(period_dropdown)
    period_unit_drpdown.select_by_visible_text(period_unit)

    driver.find_element(By.ID,"e").clear()
    driver.find_element(By.ID,"e").send_keys(period_op)

    t_value=driver.find_element(By.XPATH, '//*[@id="calc"]/div/div[8]/div/span').text
    t_value_float = float(t_value[2:].replace(",", ""))
    #print(float(t_value))
    print(float(t_value[2:].replace(",", "")))
    print(float(total_value))

    if t_value_float == float(total_value):
        sheet.cell(r, 8).value = t_value
        green = PatternFill(start_color="60b212",
                            end_color="60b212",
                            fill_type="solid")
        sheet.cell(r, 9).value = "Passed"
        sheet.cell(r, 9).fill=green
        print("Test Passed")
    else:
        sheet.cell(r, 8).value = t_value
        red = PatternFill(start_color="ff0000",
                            end_color="ff0000",
                            fill_type="solid")
        sheet.cell(r, 9).value = "Failed"
        sheet.cell(r, 9).fill=red
        print("Test Failed")
wb.save(path_of_xl)


